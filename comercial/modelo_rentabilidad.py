# © 2026 Martín Viera. Todos los derechos reservados.
"""Genera `comercial/MV_Project_Management_Analisis.xlsx`.

Un modelo de rentabilidad, no una proyección de ventas. La diferencia importa:
este archivo **no dice cuánto vas a facturar**. Dice cuánto te queda para un
conjunto de supuestos que ponés vos, y hace la aritmética —comisiones,
impuestos, ads, churn— sin equivocarse.

Todo lo que es supuesto vive en la hoja `Supuestos` y está pintado. Todo lo
demás son fórmulas de Excel de verdad (no valores calculados acá y pegados):
cambiás una celda amarilla y las 24 filas se recalculan solas. Se hizo así a
propósito — un Excel con números pegados envejece mal y nadie sabe de dónde
salió cada uno.

Qué es dato y qué es supuesto, marcado en la propia planilla:

* **Dato del producto**: los precios (US$9/mes, US$90/año) salen de
  `api/_planes.js`, que es lo que el checkout cobra de verdad.
* **Dato medible**: la comisión real de MercadoPago la da `/api/metricas`
  (`transaction_details.net_received_amount`). Hasta que haya ventas, acá va un
  supuesto y está dicho.
* **Supuesto tuyo**: clientes nuevos por mes, churn, gasto en ads, régimen
  fiscal.

No se inventó ninguna cifra de mercado ni de ventas: donde no hay fuente, hay
una celda vacía con una nota.

Se regenera con:  python comercial/modelo_rentabilidad.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parent.parent
SALIDA = RAIZ / "comercial" / "MV_Project_Management_Analisis.xlsx"

AZUL = "1F3864"
GRIS = "F2F2F2"
AMARILLO = "FFF2CC"   # celda editable
VERDE = "E2EFDA"
ROJO = "FCE4EC"

TITULO = Font(bold=True, color="FFFFFF", size=12)
NEGRITA = Font(bold=True)
CHICA = Font(size=9, italic=True, color="666666")
RELLENO_TITULO = PatternFill("solid", fgColor=AZUL)
RELLENO_EDITABLE = PatternFill("solid", fgColor=AMARILLO)
RELLENO_GRIS = PatternFill("solid", fgColor=GRIS)
RELLENO_VERDE = PatternFill("solid", fgColor=VERDE)
RELLENO_ROJO = PatternFill("solid", fgColor=ROJO)
BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def _encabezado(ws, fila, textos, ancho=None):
    for i, t in enumerate(textos, start=1):
        c = ws.cell(row=fila, column=i, value=t)
        c.font = TITULO
        c.fill = RELLENO_TITULO
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDE
    if ancho:
        for i, a in enumerate(ancho, start=1):
            ws.column_dimensions[get_column_letter(i)].width = a


def _nota(ws, fila, texto, columnas=6):
    c = ws.cell(row=fila, column=1, value=texto)
    c.font = CHICA
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=columnas)
    ws.row_dimensions[fila].height = 30


# --------------------------------------------------------------- Supuestos

def hoja_supuestos(wb):
    ws = wb.create_sheet("Supuestos")
    ws.sheet_properties.tabColor = "FFC000"

    ws["A1"] = "Supuestos — las celdas amarillas son las únicas que se tocan"
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    ws.merge_cells("A1:E1")
    _nota(ws, 2,
          "Todo lo demás en este archivo son fórmulas que dependen de acá. "
          "Cambiá un amarillo y las proyecciones de 24 meses se recalculan solas. "
          "Las filas marcadas DATO no son opinión: salen del código del producto "
          "o de una medición.", 5)

    _encabezado(ws, 4, ["Concepto", "Valor", "Unidad", "Origen", "Qué pasa si te equivocás"],
                ancho=[38, 14, 16, 14, 62])

    filas = [
        ("PRECIO Y PRODUCTO", None, None, None, None),
        ("Precio Professional mensual", 9, "USD/mes", "DATO",
         "Es lo que cobra api/_planes.js. Cambiarlo acá NO cambia el checkout."),
        ("Precio Professional anual", 90, "USD/año", "DATO",
         "12 meses al precio de 10. También de api/_planes.js."),
        ("Tipo de cambio USD→UYU", 40, "UYU/USD", "DATO",
         "El default de MP_TASA_UYU. Se cobra en pesos: si el peso se mueve, "
         "tu ingreso en USD se mueve."),
        ("% de clientes que eligen el plan anual", 0.30, "proporción", "SUPUESTO",
         "Sube el ingreso del mes 1 y baja el recurrente. Si vendés más anual de "
         "lo previsto, entra más plata antes y menos después."),

        ("VENTAS", None, None, None, None),
        ("Clientes nuevos por mes (sin ads)", 2, "clientes/mes", "SUPUESTO",
         "El supuesto MÁS sensible de todo el modelo. Con 0 no hay negocio por "
         "más que el resto esté bien."),
        ("Clientes nuevos por mes (con ads)", 6, "clientes/mes", "SUPUESTO",
         "Sólo vale si el costo por cliente de abajo es real. Poné el que midas, "
         "no el que quieras."),
        ("Churn mensual", 0.05, "proporción", "SUPUESTO",
         "5% mensual = perdés la mitad de tu base en 14 meses. Es el número que "
         "decide si esto crece o se estanca."),

        ("COSTOS VARIABLES", None, None, None, None),
        ("Comisión MercadoPago", 0.0599, "proporción", "SUPUESTO",
         "PONÉ EL TUYO. /api/metricas te da el real medido (neto que entró ÷ bruto "
         "cobrado). Varía por medio de pago y plazo de acreditación."),
        ("Costo de IA por cliente activo", 0.50, "USD/mes", "SUPUESTO",
         "Cupo de 1000 consultas/mes en el plan Professional. El motor de reglas "
         "no gasta nada; esto es sólo el copiloto."),
        ("Gasto en ads por mes", 200, "USD/mes", "SUPUESTO",
         "Escenario 'con ads'. El escenario 'sin ads' lo ignora."),
        ("Costo de adquisición por cliente (CAC)", 50, "USD/cliente", "SUPUESTO",
         "Si CAC > lo que deja un cliente en toda su vida, cada venta te hace "
         "perder plata. La planilla lo calcula abajo."),

        ("COSTOS FIJOS", None, None, None, None),
        ("Dominio web", 1.5, "USD/mes", "OBLIGATORIO",
         "~USD 15-20/año. Es el único gasto de plataforma inevitable."),
        ("Hosting (Vercel)", 0, "USD/mes", "SUPUESTO",
         "El plan Hobby es gratis pero su licencia es para uso NO comercial. "
         "Vender por ahí exige Pro. Ver la hoja 'Plataformas'."),
        ("Otros fijos (contador, etc.)", 0, "USD/mes", "SUPUESTO",
         "Un contador en Uruguay no es opcional si facturás. Poné lo que te cotice."),

        ("IMPUESTOS (Uruguay)", None, None, None, None),
        ("IVA sobre ventas locales", 0.22, "proporción", "TASA VIGENTE",
         "22% es la tasa básica. Ojo: la exportación de servicios (cliente NO "
         "residente) tiene tratamiento distinto — ver hoja 'Impuestos UY'."),
        ("% de ventas a clientes uruguayos", 0.50, "proporción", "SUPUESTO",
         "Lo que NO es local puede calificar como exportación de servicios. "
         "Cambia mucho la cuenta."),
        ("IRAE sobre la ganancia", 0.25, "proporción", "TASA VIGENTE",
         "25% sobre la renta neta fiscal en el régimen general."),
        ("¿Régimen simplificado? (1=sí, 0=no)", 0, "1/0", "SUPUESTO",
         "Con 1, el modelo ignora IRAE e IVA y usa el costo fijo de abajo. "
         "CONFIRMALO CON UN CONTADOR antes de usarlo."),
        ("Costo fijo del régimen simplificado", 0, "USD/mes", "SUPUESTO",
         "Monotributo / Literal E: aporte fijo mensual. Poné el vigente."),
    ]

    fila = 5
    for concepto, valor, unidad, origen, riesgo in filas:
        if valor is None:  # subtítulo de sección
            c = ws.cell(row=fila, column=1, value=concepto)
            c.font = NEGRITA
            for col in range(1, 6):
                ws.cell(row=fila, column=col).fill = RELLENO_GRIS
            fila += 1
            continue
        ws.cell(row=fila, column=1, value=concepto).border = BORDE
        cv = ws.cell(row=fila, column=2, value=valor)
        cv.border = BORDE
        if origen == "SUPUESTO":
            cv.fill = RELLENO_EDITABLE
        if unidad == "proporción":
            cv.number_format = "0.00%"
        ws.cell(row=fila, column=3, value=unidad).border = BORDE
        co = ws.cell(row=fila, column=4, value=origen)
        co.border = BORDE
        co.font = NEGRITA if origen in ("DATO", "OBLIGATORIO") else Font()
        cr = ws.cell(row=fila, column=5, value=riesgo)
        cr.border = BORDE
        cr.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[fila].height = 30
        fila += 1

    # Nombres para que las fórmulas de las otras hojas se lean.
    nombres = {}
    for r in range(5, fila):
        etiqueta = ws.cell(row=r, column=1).value
        if etiqueta and ws.cell(row=r, column=2).value is not None:
            nombres[etiqueta] = f"Supuestos!$B${r}"
    return ws, nombres


# -------------------------------------------------------------- Proyección

def hoja_proyeccion(wb, n, con_ads):
    titulo = "Proyección con ads" if con_ads else "Proyección sin ads"
    ws = wb.create_sheet(titulo)
    ws.sheet_properties.tabColor = "70AD47" if not con_ads else "ED7D31"

    ws["A1"] = f"{titulo} — 24 meses"
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    ws.merge_cells("A1:L1")
    _nota(ws, 2,
          "Cada fila es un mes. Las columnas son fórmulas: no hay ningún número "
          "escrito a mano acá. Los meses 1, 3, 6, 9, 12, 18 y 24 están resaltados "
          "porque son los que pediste. 'Acumulado' es lo que te queda en el "
          "bolsillo desde el mes 1 — mientras sea negativo, estás poniendo plata.", 12)

    cabeceras = ["Mes", "Clientes activos", "Altas", "Bajas (churn)",
                 "Ingreso bruto USD", "Comisión MP", "Costo IA", "Ads",
                 "Fijos", "Impuestos", "Neto del mes", "Acumulado"]
    _encabezado(ws, 4, cabeceras, ancho=[7, 15, 9, 13, 17, 12, 11, 10, 10, 12, 14, 14])

    nuevos = n["Clientes nuevos por mes (con ads)"] if con_ads \
        else n["Clientes nuevos por mes (sin ads)"]
    ads = n["Gasto en ads por mes"] if con_ads else None

    hitos = {1, 3, 6, 9, 12, 18, 24}
    primera = 5
    for m in range(1, 25):
        f = primera + m - 1
        ws.cell(row=f, column=1, value=m).font = NEGRITA

        # Clientes activos: los del mes anterior menos churn, más las altas.
        if m == 1:
            ws.cell(row=f, column=2, value=f"={nuevos}")
        else:
            ws.cell(row=f, column=2,
                    value=f"=ROUND(B{f - 1}*(1-{n['Churn mensual']})+{nuevos},0)")
        ws.cell(row=f, column=3, value=f"={nuevos}")
        ws.cell(row=f, column=4,
                value="=0" if m == 1 else f"=ROUND(B{f - 1}*{n['Churn mensual']},1)")

        # Ingreso: parte anual cobrada de una vez a los que entran, más la
        # mensualidad de los que no eligieron anual.
        anual = n["% de clientes que eligen el plan anual"]
        ws.cell(row=f, column=5, value=(
            f"=C{f}*{anual}*{n['Precio Professional anual']}"
            f"+B{f}*(1-{anual})*{n['Precio Professional mensual']}"))

        ws.cell(row=f, column=6, value=f"=E{f}*{n['Comisión MercadoPago']}")
        ws.cell(row=f, column=7, value=f"=B{f}*{n['Costo de IA por cliente activo']}")
        ws.cell(row=f, column=8, value=f"={ads}" if con_ads else "=0")
        ws.cell(row=f, column=9, value=(
            f"={n['Dominio web']}+{n['Hosting (Vercel)']}+{n['Otros fijos (contador, etc.)']}"))

        # Impuestos: o el régimen simplificado (costo fijo), o IVA sobre la
        # parte local más IRAE sobre la ganancia. La ganancia puede ser
        # negativa; ahí IRAE es 0 y no un crédito, que sería optimista.
        simplificado = n["¿Régimen simplificado? (1=sí, 0=no)"]
        iva = (f"E{f}*{n['% de ventas a clientes uruguayos']}"
               f"*{n['IVA sobre ventas locales']}")
        base = f"(E{f}-F{f}-G{f}-H{f}-I{f}-{iva})"
        ws.cell(row=f, column=10, value=(
            f"=IF({simplificado}=1,{n['Costo fijo del régimen simplificado']},"
            f"{iva}+MAX(0,{base})*{n['IRAE sobre la ganancia']})"))

        ws.cell(row=f, column=11, value=f"=E{f}-F{f}-G{f}-H{f}-I{f}-J{f}")
        ws.cell(row=f, column=12,
                value=f"=K{f}" if m == 1 else f"=L{f - 1}+K{f}")

        for col in range(1, 13):
            c = ws.cell(row=f, column=col)
            c.border = BORDE
            if col >= 5:
                c.number_format = '#,##0.00'
            if m in hitos:
                c.fill = RELLENO_VERDE

    ultima = primera + 23
    f = ultima + 2
    ws.cell(row=f, column=1, value="Mes en que el acumulado se vuelve positivo:").font = NEGRITA
    ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=3)
    ws.cell(row=f, column=4, value=(
        f'=IFERROR(MATCH(TRUE,INDEX(L{primera}:L{ultima}>0,0),0),"no en 24 meses")'))
    ws.cell(row=f, column=4).font = NEGRITA

    f += 1
    ws.cell(row=f, column=1, value="Valor de un cliente en toda su vida (LTV):").font = NEGRITA
    ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=3)
    # LTV = margen mensual / churn. Es la fórmula estándar y sale de la
    # planilla, no de un número pegado.
    ws.cell(row=f, column=4, value=(
        f"=({n['Precio Professional mensual']}*(1-{n['Comisión MercadoPago']})"
        f"-{n['Costo de IA por cliente activo']})/{n['Churn mensual']}"))
    ws.cell(row=f, column=4).number_format = '#,##0.00'

    if con_ads:
        f += 1
        c = ws.cell(row=f, column=1, value="¿El CAC se paga? (LTV − CAC):")
        c.font = NEGRITA
        ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=3)
        ws.cell(row=f, column=4,
                value=f"=D{f - 1}-{n['Costo de adquisición por cliente (CAC)']}")
        ws.cell(row=f, column=4).number_format = '#,##0.00'
        ws.cell(row=f, column=5, value=(
            "Si esto da negativo, cada peso de ads te hace perder plata: "
            "no escales, arreglá el precio o el churn primero."))
        ws.cell(row=f, column=5).font = CHICA

    ws.freeze_panes = "A5"
    return ws


# ---------------------------------------------------------- Impuestos UY

def hoja_impuestos(wb):
    ws = wb.create_sheet("Impuestos UY")
    ws["A1"] = "Impuestos en Uruguay — estructura, no asesoramiento"
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    ws.merge_cells("A1:D1")
    _nota(ws, 2,
          "Esto ordena las preguntas que le tenés que hacer a un contador; no las "
          "responde. Las tasas son las generales y son estables. Lo que NO está "
          "acá son los topes de facturación de cada régimen, que se fijan en "
          "Unidades Indexadas y se actualizan todos los años: poner un número "
          "hoy sería darte un dato que vence. Preguntalo y anotalo.", 4)

    _encabezado(ws, 4, ["Tema", "Qué aplica", "Por qué te importa acá",
                        "Qué preguntar"], ancho=[26, 26, 56, 56])

    filas = [
        ("IVA — venta local", "Tasa básica 22%",
         "Si le vendés a una empresa uruguaya, la venta lleva IVA. No es tuyo: "
         "lo cobrás y lo volcás.",
         "¿Tengo que emitir e-factura desde la primera venta? ¿Qué software de "
         "facturación electrónica me sirve?"),
        ("IVA — exportación de servicios", "Tratamiento distinto al local",
         "Un cliente en Argentina, México o España puede no llevar IVA. Con "
         "mitad de tus ventas afuera, esto cambia el resultado más que cualquier "
         "otra línea de la planilla.",
         "Mi caso (software vendido online a no residentes) ¿califica como "
         "exportación de servicios? ¿Qué necesito documentar para probarlo?"),
        ("IRAE", "25% sobre la renta neta fiscal",
         "Se paga sobre la GANANCIA, no sobre la venta. Los meses en rojo no "
         "pagan IRAE — la planilla ya lo contempla con un MAX(0;...).",
         "¿Puedo deducir el gasto en ads, el hosting y mis horas? ¿Hay anticipos "
         "mensuales?"),
        ("Regímenes simplificados", "Monotributo · IRAE Literal E",
         "Para volumen chico pueden salir bastante más baratos que el régimen "
         "general. Tienen tope de facturación: el día que lo pasás, cambiás de "
         "régimen y la cuenta se rehace entera.",
         "¿Cuál me corresponde hoy? ¿Cuál es el tope vigente? ¿Qué pasa el mes "
         "que lo supero?"),
        ("BPS", "Aportes personales",
         "Si facturás como unipersonal, aportás igual que cualquier independiente. "
         "Es un costo fijo mensual que existe aunque no vendas nada.",
         "¿Cuánto aporto por mes en mi situación, vendiendo o no?"),
        ("Retenciones del exterior", "Varía por país del cliente",
         "Algunos países retienen impuesto en origen sobre pagos al exterior. "
         "Cobrás menos de lo facturado y no es culpa de MercadoPago.",
         "¿Uruguay tiene convenio para evitar doble imposición con los países "
         "donde voy a vender?"),
    ]
    for i, (a, b, c, d) in enumerate(filas, start=5):
        for col, val in enumerate((a, b, c, d), start=1):
            cel = ws.cell(row=i, column=col, value=val)
            cel.border = BORDE
            cel.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 60

    f = len(filas) + 6
    c = ws.cell(row=f, column=1, value=(
        "No soy contador y esto no es asesoramiento fiscal. Las tasas generales "
        "(IVA 22%, IRAE 25%) son públicas y estables; el régimen que te conviene, "
        "los topes vigentes y si tu venta califica como exportación de servicios "
        "dependen de tu situación concreta y cambian. Una consulta con un contador "
        "uruguayo cuesta mucho menos que elegir mal el régimen el primer año."))
    c.font = Font(bold=True, color="9C0006")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=4)
    ws.row_dimensions[f].height = 60
    for col in range(1, 5):
        ws.cell(row=f, column=col).fill = RELLENO_ROJO


# ---------------------------------------------------------- Plataformas

def hoja_plataformas(wb):
    ws = wb.create_sheet("Plataformas")
    ws["A1"] = "Qué es obligatorio pagar y qué no"
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    ws.merge_cells("A1:E1")
    _nota(ws, 2,
          "Pediste no pagar nada salvo el dominio. Se puede casi entero. Las dos "
          "excepciones están marcadas en rojo y las dos son de licencia, no de "
          "capacidad técnica: el plan gratis alcanza de sobra para el volumen, "
          "pero su contrato no permite usarlo para vender.", 5)

    _encabezado(ws, 4, ["Plataforma", "Para qué se usa acá", "Plan gratis",
                        "¿Obligatorio pagar?", "Detalle"],
                ancho=[20, 30, 26, 20, 60])

    filas = [
        ("Dominio web", "La dirección de la landing", "No existe gratis",
         "SÍ — el único",
         "~USD 15-20/año según extensión. Es el gasto que aceptaste."),
        ("Vercel", "Landing + funciones de pago + Blob",
         "Hobby: sí, y sobra para el volumen",
         "SÍ, por licencia",
         "El plan Hobby es para uso NO COMERCIAL. Un sitio que cobra es "
         "comercial. Técnicamente andaría; el problema es el contrato, y el "
         "riesgo es que te suspendan la cuenta justo cuando estás vendiendo. "
         "Verificá los términos vigentes antes de decidir."),
        ("GitHub", "Código, CI y compilación de instaladores",
         "Sí. Repo público: Actions ilimitado. Privado: 2.000 min/mes",
         "NO (con una salvedad)",
         "Los runners de Windows consumen el DOBLE de minutos. Tres builds de "
         "instalador por push a main. Con el repo privado eso come cuota rápido "
         "— por eso los workflows ya tienen `concurrency` para no compilar dos "
         "veces lo mismo."),
        ("MercadoPago", "Cobrar", "Sin costo fijo",
         "NO, pero cobra comisión",
         "No pagás mensualidad: se queda un porcentaje de cada venta. El real lo "
         "medís en /api/metricas y lo ponés en Supuestos."),
        ("Vercel Blob", "Instalador descargable, canjes, descargas",
         "Sí, dentro del plan",
         "NO",
         "Ya está en el código. Falta el token: sin él, /api/download-installer "
         "responde 503 y el botón de la landing no baja nada."),
        ("Claude / OpenAI / Gemini", "El copiloto con IA (opcional)",
         "Hay créditos iniciales",
         "NO",
         "El motor de reglas funciona sin ninguna clave. La IA es aditiva: si no "
         "hay clave configurada, el producto anda igual."),
        ("Certificado SSL", "https", "Sí, incluido en Vercel", "NO",
         "Automático. No hay que hacer nada."),
    ]
    for i, fila in enumerate(filas, start=5):
        for col, val in enumerate(fila, start=1):
            cel = ws.cell(row=i, column=col, value=val)
            cel.border = BORDE
            cel.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 4 and str(val).startswith("SÍ"):
                cel.fill = RELLENO_ROJO
                cel.font = NEGRITA
            elif col == 4:
                cel.fill = RELLENO_VERDE
        ws.row_dimensions[i].height = 55


# ---------------------------------------------------------- Competencia

def hoja_competencia(wb):
    ws = wb.create_sheet("Competencia")
    ws["A1"] = "Contra quién competís, por mercado"
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    ws.merge_cells("A1:E1")
    _nota(ws, 2,
          "Sin precios de la competencia a propósito: cambian seguido y un número "
          "viejo acá sería peor que ninguno. Lo que sí es estable es el "
          "POSICIONAMIENTO, que es lo que decide si te eligen. Antes de una "
          "reunión de venta, mirá el precio del día en su web.", 5)

    _encabezado(ws, 4, ["Mercado", "Contra quién", "Su fuerza", "Tu ventaja real",
                        "Dónde te ganan"], ancho=[16, 30, 44, 52, 52])

    filas = [
        ("Uruguay", "Excel y Google Sheets",
         "Gratis, todos lo saben usar, cero fricción para empezar.",
         "El importador perfila el Excel que ya tienen y sugiere estructura. No "
         "les pedís que abandonen su planilla el día uno.",
         "Nada le gana a 'ya lo tengo y es gratis'. La venta es contra la "
         "inercia, no contra un producto."),
        ("Uruguay", "Consultoras locales de PMO",
         "Relación personal, presencia, factura local, hablan tu idioma.",
         "Precio por debajo de una hora de consultoría, y el producto queda "
         "corriendo cuando la consultoría termina.",
         "Confianza y respaldo. Un gerente le compra a alguien que puede ir a la "
         "oficina."),
        ("Uruguay", "Odoo / ERPs con módulo de proyectos",
         "Ya instalados en la empresa; el módulo 'viene incluido'.",
         "Salud de portafolio en 6 dimensiones y grafo de bloqueos: eso un módulo "
         "de proyectos de ERP no lo hace.",
         "Si ya pagaron el ERP, sumar un módulo es una decisión y comprarte a vos "
         "son tres."),
        ("LATAM", "Monday, Asana, ClickUp, Trello",
         "Marca, marketing enorme, integraciones con todo, apps móviles maduras.",
         "Precio en dólares por instalación, no por asiento. Corre local: el dato "
         "no sale de la empresa. Trilingüe de fábrica.",
         "Todo lo demás: app móvil, ecosistema, soporte 24/7, gente que ya sabe "
         "usarlos."),
        ("LATAM", "Jira / Atlassian",
         "Estándar de facto en equipos de software.",
         "No apuntás a equipos de software: apuntás a portafolios de proyectos "
         "con gobernanza, PMBOK y dependencias. Es otro comprador.",
         "Si el comprador es un equipo técnico, Jira ya ganó antes de que "
         "entres."),
        ("Mundo", "Smartsheet, Wrike, MS Project, Planview",
         "Venta enterprise, cumplimiento, certificaciones, integradores.",
         "Instalación en la PC del cliente sin enviar datos a la nube ajena — "
         "argumento fuerte donde el dato es sensible.",
         "Sin SOC 2 ni ISO 27001 no entrás a una compra corporativa grande, por "
         "más que el producto sea bueno."),
        ("Mundo", "Notion, Obsidian y armados caseros",
         "Flexibles, baratos, comunidad enorme.",
         "El motor de reglas calcula salud y prioridad; una plantilla de Notion "
         "no calcula nada, sólo ordena.",
         "Flexibilidad. Se adaptan a cualquier flujo; vos imponés uno."),
    ]
    for i, fila in enumerate(filas, start=5):
        for col, val in enumerate(fila, start=1):
            cel = ws.cell(row=i, column=col, value=val)
            cel.border = BORDE
            cel.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 62

    f = len(filas) + 6
    c = ws.cell(row=f, column=1, value=(
        "Lo que este cuadro dice si se lee entero: no tenés una ventaja de "
        "producto que un competidor grande no pueda copiar en un trimestre. "
        "Tenés dos ventajas de POSICIÓN —corre local sin mandar datos afuera, y "
        "el precio no es por asiento— y una de mercado: nadie grande le vende en "
        "serio a una PyME uruguaya de 10 personas. Eso es un nicho defendible, "
        "no un mercado grande. Está bien: un nicho defendible con 50 clientes "
        "pagos es un negocio."))
    c.font = NEGRITA
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=5)
    ws.row_dimensions[f].height = 75
    for col in range(1, 6):
        ws.cell(row=f, column=col).fill = RELLENO_GRIS


# --------------------------------------------------------- Calificación

def hoja_calificacion(wb):
    ws = wb.create_sheet("Calificación")
    ws.sheet_properties.tabColor = "C00000"
    ws["A1"] = "Calificación por etapa — con la evidencia de cada nota"
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    ws.merge_cells("A1:D1")
    _nota(ws, 2,
          "Cada nota sale de algo verificado en el código o medido corriendo el "
          "programa, no de una impresión. Donde no se pudo verificar, dice que no "
          "se pudo.", 4)

    _encabezado(ws, 4, ["Etapa", "Nota", "En qué se apoya", "Qué la subiría"],
                ancho=[26, 8, 72, 62])

    filas = [
        ("Código y mantenibilidad", 8,
         "857 tests que corren en CI con ruff; separación motor/UI/API real (el "
         "motor se importa sin Streamlit); los comentarios explican POR QUÉ, no "
         "qué. Varios tests documentan el bug que los originó.",
         "Cobertura medida (hoy no hay número). Type hints completos en el motor."),
        ("Seguridad", 7,
         "Firma asimétrica Ed25519 (antes era HMAC con secreto que el propio "
         "cliente se generaba: cualquiera se emitía licencia enterprise); pago "
         "verificado contra MercadoPago y no contra el query string; plan "
         "validado en la emisión; canje idempotente; rate limiting; CORS cerrado; "
         "comparaciones timing-safe; fuentes compiladas a .pyd en el instalador.",
         "Repo en privado (hoy es público con los .exe adentro). Rotar la clave "
         "—ya automatizado— y que /api/estado-licencias diga ok:true."),
        ("Funcionalidad", 7,
         "Motor de reglas completo (salud 6D, dependencias, backlog, PMBOK); tres "
         "demos que cargan sin error (sintética, portafolio UK, ClinicalTrials); "
         "API de BI + conectores Power BI/Tableau/Fabric + servidor MCP.",
         "Probarlo en una PC con Windows de verdad. Hoy los instaladores se "
         "compilan en CI y nadie los ejecutó."),
        ("Diseño de interfaz", 6,
         "Material Symbols nativos con la fuente embebida (funciona sin "
         "internet), trilingüe ES/EN/PT con test de paridad, semáforos RAG "
         "consistentes.",
         "Es Streamlit: nunca va a verse como un producto a medida. Subir de acá "
         "exige otra capa de UI, que es un proyecto en sí."),
        ("Web / landing", 7,
         "Trilingüe, CSP estricta y headers de seguridad completos, sin build ni "
         "dependencias.",
         "El botón de descarga responde 503: falta BLOB_READ_WRITE_TOKEN. Es lo "
         "primero que ve un interesado."),
        ("Cobro y licencias", 6,
         "La cadena está entera y probada cruzando lenguajes: Node emite como "
         "Vercel, Python verifica como la PC del cliente. Medido con servidores "
         "corriendo: 402 antes de pagar, 200 después.",
         "Falta MVPM_LICENSE_PRIVATE_KEY en Vercel — hoy quien pague recibe un "
         "500. Ya hay un workflow que lo resuelve con un botón. Y falta una "
         "compra real de punta a punta."),
        ("Rentabilidad realizada", 2,
         "Cero ventas. No es una opinión sobre el potencial: es que el circuito "
         "de cobro nunca se completó una vez.",
         "Una venta real. Con eso esta nota se mueve más que con seis meses de "
         "desarrollo."),
        ("GENERAL", 6.5,
         "Un producto con ingeniería bastante por encima de su madurez comercial. "
         "El código está mejor cuidado que el de mucho software que ya factura; "
         "lo que falta no es técnico, es haber vendido una vez.",
         "Cargar la clave, publicar el instalador, hacer una compra de prueba con "
         "tu propia tarjeta y mirar /api/metricas."),
    ]
    for i, (etapa, nota, apoyo, subir) in enumerate(filas, start=5):
        ws.cell(row=i, column=1, value=etapa).border = BORDE
        c = ws.cell(row=i, column=2, value=nota)
        c.border = BORDE
        c.font = Font(bold=True, size=13)
        c.alignment = Alignment(horizontal="center")
        c.fill = RELLENO_VERDE if nota >= 7 else (
            RELLENO_ROJO if nota < 5 else RELLENO_EDITABLE)
        for col, val in ((3, apoyo), (4, subir)):
            cel = ws.cell(row=i, column=col, value=val)
            cel.border = BORDE
            cel.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 76
        if etapa == "GENERAL":
            ws.cell(row=i, column=1).font = Font(bold=True, size=12)


def hoja_palancas(wb, n):
    """Sensibilidad al precio y al churn.

    Es la hoja más útil de todo el archivo y la que menos se parece a lo que se
    pidió, así que conviene decir por qué existe. Corriendo el modelo con los
    supuestos de arriba, a 24 meses y sin ads, quedan poco más de USD 2.000
    acumulados. Eso no es un negocio que reemplace un sueldo, y la causa no es
    el producto: es que USD 9 por mes por INSTALACIÓN —no por asiento— es un
    precio de app de consumo puesto a un software que le vendés a una empresa.

    Esta hoja mide cuánto mueve cada palanca, para que la decisión de precio se
    tome mirando un número y no una sensación.
    """
    ws = wb.create_sheet("Palancas")
    ws.sheet_properties.tabColor = "7030A0"
    ws["A1"] = "Qué mueve más la aguja: el precio o vender más"
    ws["A1"].font = Font(bold=True, size=14, color=AZUL)
    ws.merge_cells("A1:F1")
    _nota(ws, 2,
          "LTV = lo que deja un cliente en toda su vida = margen mensual ÷ churn. "
          "Es la fórmula estándar y acá está viva: cambiá los supuestos y esto se "
          "mueve. Compará las filas entre sí, no mires el número absoluto.", 6)

    _encabezado(ws, 4, ["Escenario", "Precio USD/mes", "Churn mensual",
                        "LTV por cliente", "vs. hoy", "Lectura"],
                ancho=[30, 15, 15, 16, 12, 60])

    escenarios = [
        ("HOY", None, None, "El punto de partida."),
        ("Precio ×2 (USD 19)", 19, None,
         "Duplicar el precio duplica el LTV sin vender un cliente más. Es la "
         "palanca más barata que tenés: no requiere ads, ni tiempo, ni código."),
        ("Precio ×3 (USD 29)", 29, None,
         "USD 29/mes por una empresa entera sigue estando MUY por debajo de "
         "Monday o Asana, que cobran por persona."),
        ("Churn a la mitad (2,5%)", None, 0.025,
         "Que el cliente se quede el doble de tiempo también duplica el LTV. Más "
         "difícil que subir el precio, pero se sostiene solo."),
        ("Precio ×3 y churn a la mitad", 29, 0.025,
         "Las dos juntas se multiplican, no se suman."),
    ]

    precio_base = n["Precio Professional mensual"]
    churn_base = n["Churn mensual"]
    com = n["Comisión MercadoPago"]
    ia = n["Costo de IA por cliente activo"]

    for i, (nombre, precio, churn, lectura) in enumerate(escenarios, start=5):
        ws.cell(row=i, column=1, value=nombre).font = NEGRITA
        ws.cell(row=i, column=2,
                value=f"={precio_base}" if precio is None else precio)
        cc = ws.cell(row=i, column=3, value=f"={churn_base}" if churn is None
                     else churn)
        cc.number_format = "0.00%"
        ws.cell(row=i, column=4, value=f"=(B{i}*(1-{com})-{ia})/C{i}").number_format = '#,##0'
        cv = ws.cell(row=i, column=5, value="" if i == 5 else f"=D{i}/$D$5")
        cv.number_format = '0.0"×"'
        cv.font = NEGRITA
        cl = ws.cell(row=i, column=6, value=lectura)
        cl.alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(1, 7):
            ws.cell(row=i, column=col).border = BORDE
        ws.row_dimensions[i].height = 46

    f = len(escenarios) + 6
    c = ws.cell(row=f, column=1, value=(
        "La conclusión incómoda: con los supuestos de la hoja Supuestos, a 24 "
        "meses y sin ads quedan poco más de USD 2.000 acumulados. El cuello de "
        "botella no es el producto ni el marketing — es el precio. USD 9/mes por "
        "una instalación completa es lo que cobra una app de teléfono, no un "
        "software de gestión de portafolios que le vendés a una empresa con "
        "PMBOK, grafo de dependencias y conectores de BI. Antes de gastar un "
        "dólar en ads, probá cobrar tres veces más: si con USD 29 no te compran, "
        "con USD 9 tampoco te iban a comprar por mucho tiempo."))
    c.font = NEGRITA
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=6)
    ws.row_dimensions[f].height = 90
    for col in range(1, 7):
        ws.cell(row=f, column=col).fill = RELLENO_EDITABLE


def main():
    wb = Workbook()
    wb.remove(wb.active)
    _, nombres = hoja_supuestos(wb)
    hoja_proyeccion(wb, nombres, con_ads=False)
    hoja_proyeccion(wb, nombres, con_ads=True)
    hoja_impuestos(wb)
    hoja_plataformas(wb)
    hoja_competencia(wb)
    hoja_palancas(wb, nombres)
    hoja_calificacion(wb)
    SALIDA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SALIDA)
    print(f"Generado: {SALIDA.relative_to(RAIZ)} ({SALIDA.stat().st_size / 1024:.0f} KB)")
    print(f"Hojas: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
